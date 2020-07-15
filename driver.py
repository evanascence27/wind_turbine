#driver code
from PredictionModels import run_models
from openpyxl import load_workbook
from tkinter import *
import tkinter as tk
from tkinter import messagebox as mb
import pandas as pd
from PIL import ImageTk, Image
from tkcalendar import DateEntry
from datetime import date

def excel(sheet): 
    # at particular location 
    sheet.cell(row=1, column=1).value = "Date/Time"
    sheet.cell(row=1, column=2).value = "LV ActivePower (kW)"
    sheet.cell(row=1, column=3).value = "Wind Speed (m/s)"
    sheet.cell(row=1, column=4).value = "Theoretical_Power_Curve (KWh)"
    sheet.cell(row=1, column=5).value = "Wind Direction (°)"

def insert(sheet, current_row, active_pow, wind_speed ,wind_dir, theo_pow, datetime):
    #current_column = sheet.max_column
    sheet.cell(row = current_row, column = 1).value = datetime
    sheet.cell(row = current_row, column = 2).value = active_pow
    sheet.cell(row = current_row, column = 3).value = wind_speed
    sheet.cell(row = current_row, column = 4).value = theo_pow
    sheet.cell(row = current_row, column = 5).value = wind_dir

def entry_destroy():
    txt_ws.delete(first=0, last=100)
    txt_wd.delete(first=0, last=100)
    txt_theo_pow.delete(first=0, last=100)
    dt.delete(first=0, last=100)
    hrs.delete(first=0, last=100)
    mins.delete(first=0, last=100)

def dn_clicked():
    mb.showinfo('Wind Turbine Data Submission','Thank you!')
    window.destroy()

def btn_clicked():
    d = dt.get().replace('/', ' ')
    h = int(hrs.get())
    m = int(mins.get())
    t = str(h) + ":" + str(m)
    datetime = str(d + " " + t)
    wind_speed = float(txt_ws.get())
    wind_dir = float(txt_wd.get())
    theo_pow = float(txt_theo_pow.get())
    if (datetime and wind_speed and wind_dir and h >= 0 and h <= 24 and m >= 0 and m < 60):
        #the user entered data in the mandatory entry: proceed to next step
        mb.showinfo('Wind Turbine Data Submission','Press OK to start calculating your data....\n')
        #insert new data
        active_pow = 0
        insert(sheet, sheet.max_row+1, active_pow, wind_speed, wind_dir, theo_pow, datetime)
        #saving the dataframe 
        wb.save('dataset.xlsx')
        data_xls = pd.read_excel('dataset.xlsx', index_col = None)
        data_xls.to_csv('dataset.csv', encoding = 'utf-8', index = False)
        #model call
        model, ap_pred =  run_models(path)
        #predict power
        active_pow = round(ap_pred[len(ap_pred)-1], 7)
        insert(sheet, sheet.max_row, active_pow, wind_speed, wind_dir, theo_pow, datetime)
        #saving the dataframe 
        wb.save('dataset.xlsx')
        data_xls = pd.read_excel('dataset.xlsx', index_col = None)
        data_xls.to_csv('dataset.csv', encoding = 'utf-8', index = False)
        #displaying predicted value
        ap_out = tk.Message(window, text = str(active_pow), width = 85, background = "pale turquoise", relief = "sunken")
        ap_out.place(x = 210, y = 550)
        window.mainloop()
    else:
        #the mandatory field is empty
        mb.showwarning('Wind Turbine Data Submission','Some fields are empty or filled incorrectly!')
    #entry_destroy()    

def cncl_clicked():
    mb.showinfo('Wind Turbine Data Submission','Your submission will be cancelled.')
    window.destroy()

#load workbook
def load():
    wb = load_workbook("dataset.xlsx")  
    sheet = wb.active
    return wb, sheet

#window
def create_window():
    #main window
    window = Tk()
    window.title("Wind Turbine Data")
    window.geometry('1250x680')
    window.iconbitmap(default = "logo.ico")
    window.config(background = "white")
    img = ImageTk.PhotoImage(Image.open('bg1.png').resize((40, 770), Image.ANTIALIAS))
    pic3 = tk.Label(window, image = img, borderwidth = 0)
    pic3.image = img
    pic3.place(x = 0, y = 0)
    img = ImageTk.PhotoImage(Image.open('bg2.png').resize((225, 225), Image.ANTIALIAS))
    pic4 = tk.Label(window, image = img, borderwidth = 0)
    pic4.image = img
    pic4.place(x = 181, y = 480)
    return window

def create_labels(window, ts_pic, logo_pic):

     #heading
    lbl = tk.Label(window, text = "WIND TURBINE DATA INPUT", font = ("Helvetica", 14, "bold"), background = "white")
    lbl.place(x = 50, y = 80)
    #logo
    img = ImageTk.PhotoImage(Image.open(logo_pic).resize((220, 100), Image.ANTIALIAS))
    pic1 = tk.Label(window, image = img, borderwidth = 0)
    pic1.image = img
    pic1.place(x = 70, y = 120)
    #time series pics
    lbl = tk.Label(window, text = "TIME SERIES FORECAST & SEASONALITY", font = ("Helvetica", 14, "bold"), background = "white")
    lbl.place(x = 620, y = 15)
    lbl = tk.Label(window, text = "Wind Direction", font = ("Helvetica", 13), background = "white")
    lbl.place(x = 540, y = 50)
    lbl = tk.Label(window, text = "Wind Speed", font = ("Helvetica", 13), background = "white")
    lbl.place(x = 970, y = 50)
    #wind speed
    txt = tk.Label(window, text = "Wind Speed(kmph):", font = ("Helvetica", 10), background = "white")
    txt.place(x = 50, y = 250)
    txt_ws = tk.Entry(window, width = 15, background = 'pale turquoise')
    txt_ws.place(x = 210, y = 250)
    #wind direction
    txt = tk.Label(window, text = "Wind Direction(deg):", font = ("Helvetica", 10), background = "white")
    txt.place(x = 50, y = 300)
    txt_wd = tk.Entry(window, width = 15, background = 'pale turquoise')
    txt_wd.place(x = 210, y = 300)
    #country
    txt = tk.Label(window, text = "Theoretical Power(kW):", font = ("Helvetica", 10), background = "white")
    txt.place(x = 50, y = 350)
    txt_theo_pow = tk.Entry(window, width = 15, background = 'pale turquoise')
    txt_theo_pow.place(x = 210, y = 350)
    #date
    txt = tk.Label(window, text = "Date(MM/DD/YYYY):", font = ("Helvetica", 10), background = "white")
    txt.place(x = 50, y = 400)
    dt = DateEntry(window, locale = 'en_US', date_pattern='MM/dd/yyyy', background = 'pale turquoise')
    dt.place(x = 210, y = 400)
    #time
    txt = tk.Label(window, text = "Time(hr:min):", font = ("Helvetica", 10), background = "white")
    txt.place(x = 50, y = 450)
    hrs = tk.Entry(window, width = 2, background = 'pale turquoise')
    hrs.place(x = 210, y = 450)
    txt = tk.Label(window, text = ":", font = ("Helvetica", 10), background = "white")
    txt.place(x = 230, y = 450)
    mins = tk.Entry(window, width = 2, background = 'pale turquoise')
    mins.place(x = 245, y = 450)
    #predicted output
    txt = tk.Label(window, text = "Active Power(predicted):", font = ("Helvetica", 10), background = "white")
    txt.place(x = 50, y = 550)
    #time series display
    img = ImageTk.PhotoImage(Image.open(ts_pic).resize((800, 580), Image.ANTIALIAS))
    pic2 = tk.Label(window, image = img, borderwidth = 0)
    pic2.image = img
    pic2.place(x = 400, y = 80)
    return txt_ws, txt_wd, txt_theo_pow, dt, hrs, mins

#submit or cancel
def create_sub_can():
    tk.Button(window, text = "Submit", command = btn_clicked).place(x = 210, y = 490)
    tk.Button(window, text = "Finish", command = dn_clicked).place(x = 50, y = 640)
    tk.Button(window, text = "Cancel", command = cncl_clicked).place(x = 260, y = 640)

#sequence
path = "dataset.csv"
wb, sheet = load()
excel(sheet)
window = create_window()
logo_pic = 'logo_pic.png'
ts_pic = 'ts_seasonality.png'
txt_ws, txt_wd, txt_theo_pow, dt, hrs, mins = create_labels(window, ts_pic, logo_pic)
create_sub_can()
window.mainloop()








